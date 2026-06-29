"""
Exportación a Power BI — Banco Digital Chile (IL3.1 / EP3)
==========================================================
Produce los insumos del dashboard de monitoreo:

  dashboard/powerbi_dataset.csv     → tabla de hechos (1 fila por traza) para Power BI
  dashboard/dashboard_banco.xlsx    → workbook con tablas + gráficos nativos (vista rápida)
  (la guía de armado está en dashboard/GUIA_POWERBI.md)

Uso:
    python -m observability.export_powerbi
"""

from __future__ import annotations

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from observability import metrics as M
from observability.analyze_traces import (
    anomalias_temporales, puntos_de_falla,
)

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DASH_DIR = os.path.join(_ROOT, "dashboard")
os.makedirs(DASH_DIR, exist_ok=True)

CSV_PATH = os.path.join(DASH_DIR, "powerbi_dataset.csv")
XLSX_PATH = os.path.join(DASH_DIR, "dashboard_banco.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _tabla_hechos(df: pd.DataFrame) -> pd.DataFrame:
    """Aplana las trazas a una tabla de hechos (sin columnas anidadas)."""
    cols = [
        "trace_id", "session_id", "timestamp", "fecha", "hora", "escenario",
        "variante", "intent", "modelo", "expected_tool", "n_steps", "n_tools",
        "latency_ms", "tokens_prompt", "tokens_completion", "tokens_total",
        "cost_usd", "success", "correct_tool", "error_type",
        "pii_detected", "injection_blocked", "masked_fields",
    ]
    fact = df[cols].copy()
    fact["exito_int"] = fact["success"].astype(int)
    fact["error_int"] = fact["error_type"].notna().astype(int)
    return fact


def _escribir_hoja(ws, df: pd.DataFrame, index=False):
    for r in dataframe_to_rows(df, index=index, header=True):
        ws.append(r)
    # Estilo de encabezado
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def construir_xlsx(df: pd.DataFrame):
    fact = _tabla_hechos(df)
    g = M.global_metrics(df)
    esc = M.por_escenario(df).reset_index()
    serie = M.serie_diaria(df)
    serie["fecha"] = serie["fecha"].astype(str)
    err = M.errores(df)
    cons_tabla, cons = M.consistencia(df)
    sec = M.seguridad(df)
    fallas = puntos_de_falla(df)

    wb = Workbook()

    # 1) KPIs
    ws = wb.active
    ws.title = "KPIs"
    ws.append(["KPI", "Valor"])
    kpis = [
        ("Interacciones totales", g["interacciones_totales"]),
        ("Precisión (éxito)", f"{g['precision_exito']*100:.1f}%"),
        ("Acierto de herramienta", f"{g['acierto_herramienta']*100:.1f}%"),
        ("Tasa de error", f"{g['tasa_error']*100:.1f}%"),
        ("Consistencia media", f"{cons['consistencia_media']*100:.1f}%"),
        ("Latencia media (ms)", g["latencia_media_ms"]),
        ("Latencia p95 (ms)", g["latencia_p95_ms"]),
        ("Latencia máx (ms)", g["latencia_max_ms"]),
        ("Tokens totales", g["tokens_totales"]),
        ("Costo estimado (US$)", g["costo_total_usd"]),
        ("Intentos injection", sec["intentos_injection"]),
        ("Injection bloqueados", sec["bloqueados"]),
        ("Injection bypasses", sec["bypasses"]),
        ("Interacciones con PII", sec["interacciones_con_pii"]),
    ]
    for k, v in kpis:
        ws.append([k, v])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16

    # 2) PorEscenario (+ gráficos)
    ws_e = wb.create_sheet("PorEscenario")
    _escribir_hoja(ws_e, esc)
    n = len(esc)

    bar_lat = BarChart()
    bar_lat.type = "bar"
    bar_lat.title = "Latencia p95 por escenario (ms)"
    cats = Reference(ws_e, min_col=1, min_row=2, max_row=n + 1)
    col_p95 = list(esc.columns).index("latencia_p95_ms") + 1
    data = Reference(ws_e, min_col=col_p95, min_row=1, max_row=n + 1)
    bar_lat.add_data(data, titles_from_data=True)
    bar_lat.set_categories(cats)
    bar_lat.height, bar_lat.width = 8, 16
    ws_e.add_chart(bar_lat, "K2")

    bar_prec = BarChart()
    bar_prec.type = "bar"
    bar_prec.title = "Precisión por escenario"
    col_prec = list(esc.columns).index("precision_exito") + 1
    data2 = Reference(ws_e, min_col=col_prec, min_row=1, max_row=n + 1)
    bar_prec.add_data(data2, titles_from_data=True)
    bar_prec.set_categories(cats)
    bar_prec.height, bar_prec.width = 8, 16
    ws_e.add_chart(bar_prec, "K18")

    # 3) SerieDiaria (+ línea)
    ws_s = wb.create_sheet("SerieDiaria")
    _escribir_hoja(ws_s, serie)
    m = len(serie)
    line = LineChart()
    line.title = "Latencia p95 diaria (ms)"
    cats_s = Reference(ws_s, min_col=1, min_row=2, max_row=m + 1)
    col_lat = list(serie.columns).index("latencia_p95_ms") + 1
    data_s = Reference(ws_s, min_col=col_lat, min_row=1, max_row=m + 1)
    line.add_data(data_s, titles_from_data=True)
    line.set_categories(cats_s)
    line.height, line.width = 8, 18
    ws_s.add_chart(line, "J2")

    line2 = LineChart()
    line2.title = "Tasa de error diaria"
    col_err = list(serie.columns).index("tasa_error") + 1
    data_e = Reference(ws_s, min_col=col_err, min_row=1, max_row=m + 1)
    line2.add_data(data_e, titles_from_data=True)
    line2.set_categories(cats_s)
    line2.height, line2.width = 8, 18
    ws_s.add_chart(line2, "J18")

    # 4) Errores (+ pie)
    ws_err = wb.create_sheet("Errores")
    _escribir_hoja(ws_err, err)
    if len(err):
        pie = PieChart()
        pie.title = "Distribución de errores"
        labels = Reference(ws_err, min_col=1, min_row=2, max_row=len(err) + 1)
        data_p = Reference(ws_err, min_col=2, min_row=1, max_row=len(err) + 1)
        pie.add_data(data_p, titles_from_data=True)
        pie.set_categories(labels)
        pie.height, pie.width = 8, 12
        ws_err.add_chart(pie, "F2")

    # 5) Consistencia, PuntosFalla, Trazas
    ws_c = wb.create_sheet("Consistencia")
    _escribir_hoja(ws_c, cons_tabla)
    if not fallas.empty:
        ws_f = wb.create_sheet("PuntosFalla")
        _escribir_hoja(ws_f, fallas)
    ws_t = wb.create_sheet("Trazas")
    _escribir_hoja(ws_t, fact)

    wb.save(XLSX_PATH)
    return fact


def exportar():
    df = M.load_df()
    fact = construir_xlsx(df)
    fact.to_csv(CSV_PATH, index=False, encoding="utf-8-sig")
    return fact


if __name__ == "__main__":
    fact = exportar()
    print(f"✔ Power BI dataset : {os.path.relpath(CSV_PATH, _ROOT)}  ({len(fact)} filas)")
    print(f"✔ Workbook Excel    : {os.path.relpath(XLSX_PATH, _ROOT)}")
